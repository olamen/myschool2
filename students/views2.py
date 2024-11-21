# students/views.py
import datetime
from datetime import date
from django.forms import ValidationError
from django.http import HttpResponse
from django.views import View
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfgen import canvas
from django.conf import settings
import os
from .models import Student, Class
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from arabic_reshaper import reshape
from bidi.algorithm import get_display

# class ListStudentPDFView(View):
#     def get(self, request, *args, **kwargs):
#         # Define the HTTP response with PDF header
#         response = HttpResponse(content_type='application/pdf')
#         response['Content-Disposition'] = 'attachment; filename="liste_etudiants.pdf"'
        
#         # Create a PDF document with reduced top margin
#         doc = SimpleDocTemplate(response, pagesize=A4, topMargin=0.5 * inch, bottomMargin=0.5 * inch)
#         elements = []
        
#         # Define styles for the document
#         styles = getSampleStyleSheet()
#         title_style = styles['Heading1']
#         normal_style = styles['Normal']
        
#         # Path to the logo image
#         logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo1.png')  # Adjust the path as needed
        
#         # Prepare the school information text on the left side
#         school_info = [
#             Paragraph("Nom de l'école: École Exemple", title_style),
#             Paragraph("Adresse de l'école: 123 Rue de l'École, Ville, Pays", normal_style),
#             Paragraph("Téléphone: +123456789 | Email: ecole@example.com", normal_style),
#         ]
        
#         # Add logo image on the right side if it exists
#         if os.path.exists(logo_path):
#             logo = Image(logo_path, width=1.5 * inch, height=1.5 * inch)
#         else:
#             logo = Paragraph("مدرسة احيال المستقبل", normal_style)  # Empty paragraph if logo is missing
        
#         # Create a table with school info on the left and logo on the right
#         header_table = Table([[school_info, logo, school_info]], colWidths=[7.0 * inch, 1.0 * inch])
#         header_table.setStyle(TableStyle([
#             ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Align contents to the top
#         ]))
        
#         elements.append(header_table)
#         elements.append(Spacer(1, 10))  # Add space before the main table
        
#         # Table header for student information in French
#         data = [["Prénom", "Nom", "NNI", "Téléphone", "Date d'inscriptions", "Classe"]]
        
#         # Fetch students and populate the data list
#         students = Student.objects.all()
#         for student in students:
#             data.append([
#                 student.first_name,
#                 student.last_name,
#                 student.nni,
#                 student.mobile,
#                 student.enrollment_date.strftime('%Y-%m-%d'),
#                 student.student_class.name
#             ])
        
#         # Create table with data
#         table = Table(data)
#         table.setStyle(TableStyle([
#             ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
#             ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
#             ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
#             ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
#             ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
#             ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
#             ('GRID', (0, 0), (-1, -1), 1, colors.black),
#         ]))
        
#         elements.append(table)
        
#         # Build PDF with header and table, with page numbering callback
#         doc.build(elements, onLaterPages=self.add_page_number)
        
#         return response

#     def add_page_number(self, canvas, doc):
#         """Add page number to the bottom of each page."""
#         page_num = canvas.getPageNumber()
#         text = f"Page {page_num}"
#         canvas.setFont("Helvetica", 9)
#         canvas.drawRightString(A4[0] - inch, 0.75 * inch, text)  # Positioned 0.75 inch from the bottom right




class ListStudentPDFView(View):
    def get(self, request, *args, **kwargs):
        # Define the HTTP response with PDF header
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="liste_etudiants.pdf"'
        
        # Create a PDF document with reduced top margin
        doc = SimpleDocTemplate(response, pagesize=A4, topMargin=0.5 * inch, bottomMargin=0.5 * inch)
        elements = []
        
        # Register Arabic font
        font_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'Amiri-Regular.ttf')
        pdfmetrics.registerFont(TTFont('Amiri', font_path))
        
        # Define styles for the document and set the font
        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        title_style.fontName = 'Amiri'
        title_style.fontSize = 16
        normal_style = styles['Normal']
        normal_style.fontName = 'Amiri'
        
        # Reshape and bidi-process the Arabic text for the school name
        school_name = get_display(reshape("مدرسة أجيال المستقبل"))
        address = get_display(reshape("عنوان المدرسة: 123 شارع النصر، تفرغ زينة"))
        contact = get_display(reshape("هاتف: +222 22 24 24 20 | بريد إلكتروني: school@example.com"))
        
        # Prepare the school information text with Arabic content
        school_info = [
            Paragraph(school_name, title_style),
            Paragraph(address, normal_style),
            Paragraph(contact, normal_style),
        ]
        
        # Path to the logo image
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'logo.png')  # Adjust the path as needed
        
        # Add logo image on the right side if it exists
        if os.path.exists(logo_path):
            logo = Image(logo_path, width=1.5 * inch, height=1.5 * inch)
        else:
            logo = Paragraph("", normal_style)  # Empty paragraph if logo is missing
        
        # Create a table with school info on the left and logo on the right
        header_table = Table([[school_info, logo]], colWidths=[4.5 * inch, 1.5 * inch])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Align contents to the top
        ]))
        
        elements.append(header_table)
        elements.append(Spacer(1, 20))  # Add space before the main table
        
        # Table header for student information in Arabic
        data = [["Prénom", "Nom", "NNI", "Téléphone", "Date d'inscriptions", "Classe"]]
        
        # Fetch students and populate the data list
        students = Student.objects.all()
        for student in students:
            data.append([
                student.first_name,
                student.last_name,
                student.nni,
                student.mobile,
                student.enrollment_date.strftime('%Y-%m-%d'),
                student.student_class.name
            ])
        
        # Create table with data
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Amiri'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        elements.append(table)
        
        # Build PDF with header and table
        doc.build(elements, onLaterPages=self.add_page_number)
        
        return response

    def add_page_number(self, canvas, doc):
        """Add page number to the bottom of each page."""
        page_num = canvas.getPageNumber()
        text = f"الصفحة {page_num}"
        canvas.setFont("Amiri", 9)
        canvas.drawRightString(A4[0] - inch, 0.75 * inch, text)  # Positioned 0.75 inch from the bottom right

from django.http import HttpResponse
from openpyxl import Workbook

class GenerateExcelTemplateView(View):
    def get(self, request, *args, **kwargs):
        # Create a workbook and add a worksheet
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Student Template"

        # Define the headers for the template
        headers = ["First Name", "Last Name", "NNI", "Mobile", "Enrollment Date", "Class", "Gender", "Has Discount"]
        worksheet.append(headers)  # Add headers to the first row

        # Set column widths for better readability
        for column in worksheet.columns:
            max_length = 20  # Adjust as necessary
            column_letter = column[0].column_letter  # Get the column letter
            worksheet.column_dimensions[column_letter].width = max_length

        # Create an HTTP response with the appropriate header for Excel
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="student_template.xlsx"'

        # Save the workbook to the response
        workbook.save(response)
        return response
    
# students/views.py
from openpyxl import load_workbook
from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils.dateparse import parse_date
from django.urls import reverse

class BulkUploadStudentsView(View):
    def get(self, request, *args, **kwargs):
        """Render the file upload form."""
        return render(request, 'students/bulk_upload.html')

    def post(self, request, *args, **kwargs):
        file = request.FILES.get('excel_file')  # The uploaded file
        
        # Check if file is uploaded
        if not file:
            messages.error(request, "Please upload an Excel file.")
            return redirect('bulk_upload_students')
        
        # Load the workbook
        workbook = load_workbook(file)
        worksheet = workbook.active

        # Process each row after the header
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            # Read data from each cell in the row
            first_name = row[0]
            last_name = row[1]
            nni = row[2]
            mobile = row[3]
            class_name = row[4]
            gender = row[5]
            has_discount = bool(int(row[6])) if row[6] is not None else False  # Convert 0 or 1 to boolean

            # Debugging: Print each row's data to the console
            print(f"Row data - First Name: {first_name}, Last Name: {last_name}, NNI: {nni}, "
                  f"Mobile: {mobile}, Class: {class_name}, "
                  f"Gender: {gender}, Has Discount: {has_discount}")

            # Find the Class instance
            try:
                student_class = Class.objects.get(name=class_name)
            except Class.DoesNotExist:
                messages.error(request, f"Class '{class_name}' does not exist.")
                return redirect('bulk_upload_students')

            # Create the Student record and handle any errors
            try:
                student = Student.objects.create(
                    first_name=first_name,
                    last_name=last_name,
                    nni=nni,
                    mobile=mobile,
                    enrollment_date=date.today,
                    student_class=student_class,
                    gender=gender,
                    has_discount=has_discount
                )
                print(f"Student '{student.first_name} {student.last_name}' added successfully.")
            except Exception as e:
                # Catch any errors and show a message
                messages.error(request, f"Error adding student: {e}")
                print(f"Error adding student: {e}")
                return redirect('bulk_upload_students')

        messages.success(request, "Students uploaded successfully.")
        return redirect(reverse('students_list'))